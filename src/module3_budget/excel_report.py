"""
Module 3 — Excel Budget Recommendation Report
Generates a CMO-facing Excel workbook with 4 sheets:
  1. Executive Summary
  2. Budget Recommendation
  3. Incrementality Results
  4. Attribution Comparison
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
import os
from datetime import datetime


# ── Color palette ──────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
GREEN       = "1E8449"
LIGHT_GREEN = "D5F5E3"
RED         = "C0392B"
LIGHT_RED   = "FADBD8"
YELLOW_BG   = "FFF9C4"
WHITE       = "FFFFFF"
GRAY        = "F2F2F2"
DARK_GRAY   = "595959"


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")

def _border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ── SHEET 1: Executive Summary ─────────────────────────────

def build_executive_summary(wb, inc_df, budget_df):
    ws = wb.create_sheet("Executive Summary", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "AD CAMPAIGN ATTRIBUTION & BUDGET RECOMMENDATION"
    ws["A1"].font = _font(bold=True, color=WHITE, size=16)
    ws["A1"].fill = _fill(DARK_BLUE)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y')}   |   Total Budget: ₹1,550,000"
    ws["A2"].font = _font(italic=True, color=WHITE, size=10)
    ws["A2"].fill = _fill(MID_BLUE)
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 22

    # KPI Cards row
    ws.row_dimensions[4].height = 18
    kpis = [
        ("Total Channels", "6", MID_BLUE),
        ("Channels Tested", "6", MID_BLUE),
        ("Channels w/ Positive ROAS", str(len(inc_df[inc_df["incremental_roas"] > 0])), GREEN),
        ("Estimated Wasted Spend", f"₹{inc_df[inc_df['incremental_roas']<1]['ad_spend'].sum():,.0f}", RED),
        ("Projected Revenue Gain", f"₹{budget_df['Projected Rev Delta'].sum():,.0f}", GREEN),
    ]

    col_map = ["A", "B", "C", "D", "E"]
    ws.row_dimensions[5].height = 14
    for i, (label, value, color) in enumerate(kpis):
        col = col_map[i]
        ws.merge_cells(f"{col}5:{col}6")
        ws[f"{col}5"] = label
        ws[f"{col}5"].font = _font(bold=True, color=WHITE, size=9)
        ws[f"{col}5"].fill = _fill(color)
        ws[f"{col}5"].alignment = _center()
        ws.merge_cells(f"{col}7:{col}8")
        ws[f"{col}7"] = value
        ws[f"{col}7"].font = _font(bold=True, color=color, size=14)
        ws[f"{col}7"].fill = _fill(GRAY)
        ws[f"{col}7"].alignment = _center()
        ws.row_dimensions[7].height = 26

    # Key Findings
    ws.merge_cells("A10:G10")
    ws["A10"] = "KEY FINDINGS"
    ws["A10"].font = _font(bold=True, color=WHITE, size=12)
    ws["A10"].fill = _fill(DARK_BLUE)
    ws["A10"].alignment = _center()
    ws.row_dimensions[10].height = 22

    findings = []
    for _, row in inc_df.iterrows():
        if row["incremental_roas"] < 0 and row["significant_95"]:
            findings.append(("🔴 STOP", row["channel"],
                f"Negative ROAS {row['incremental_roas']:.2f}x — proven to hurt conversions",
                RED, LIGHT_RED))
        elif row["incremental_roas"] < 1:
            findings.append(("🟡 REDUCE", row["channel"],
                f"Sub-1x ROAS {row['incremental_roas']:.2f}x — spending more than earning",
                "7D6608", YELLOW_BG))
        elif row["incremental_roas"] >= 3:
            findings.append(("✅ INCREASE", row["channel"],
                f"Strong ROAS {row['incremental_roas']:.2f}x — high incremental value",
                GREEN, LIGHT_GREEN))

    headers = ["Action", "Channel", "Insight"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=11, column=j, value=h)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.fill = _fill(MID_BLUE)
        cell.alignment = _center()
        cell.border = _border()
    ws.row_dimensions[11].height = 18

    for r, (action, channel, insight, txt_color, bg_color) in enumerate(findings, 12):
        ws.cell(row=r, column=1, value=action).font = _font(bold=True, color=txt_color, size=10)
        ws.cell(row=r, column=1).fill = _fill(bg_color)
        ws.cell(row=r, column=1).alignment = _center()
        ws.cell(row=r, column=2, value=channel).font = _font(bold=True, size=10)
        ws.cell(row=r, column=2).alignment = _left()
        ws.cell(row=r, column=3, value=insight).font = _font(size=10)
        ws.cell(row=r, column=3).alignment = _left()
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = _border()
        ws.row_dimensions[r].height = 18

    # Column widths
    _set_col_width(ws, "A", 14)
    _set_col_width(ws, "B", 20)
    _set_col_width(ws, "C", 55)
    _set_col_width(ws, "D", 20)
    _set_col_width(ws, "E", 22)


# ── SHEET 2: Budget Recommendation ─────────────────────────

def build_budget_sheet(wb, budget_df, inc_df):
    ws = wb.create_sheet("Budget Recommendation")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "BUDGET REALLOCATION RECOMMENDATION — CMO VIEW"
    ws["A1"].font = _font(bold=True, color=WHITE, size=14)
    ws["A1"].fill = _fill(DARK_BLUE)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = "Blue = hardcoded inputs  |  Black = calculated formulas  |  Green = positive impact  |  Red = reduce/stop"
    ws["A2"].font = _font(italic=True, color=DARK_GRAY, size=9)
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 16

    headers = ["Channel", "Current Spend (₹)", "Recommended (₹)",
               "Δ Spend (₹)", "Inc. ROAS", "Current Revenue (₹)",
               "Projected Revenue (₹)", "Revenue Δ (₹)"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=j, value=h)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.fill = _fill(MID_BLUE)
        cell.alignment = _center()
        cell.border = _border()
    ws.row_dimensions[4].height = 30

    roas_map = dict(zip(inc_df["channel"], inc_df["incremental_roas"]))

    for r, (_, row) in enumerate(budget_df.iterrows(), 5):
        channel  = row["Channel"]
        curr     = row["Current Spend (₹)"]
        rec      = row["Recommended (₹)"]
        roas_val = roas_map.get(channel, 0)
        delta    = rec - curr

        # Determine row color
        if roas_val < 0:
            row_bg = LIGHT_RED
        elif roas_val < 1:
            row_bg = YELLOW_BG
        elif roas_val >= 3:
            row_bg = LIGHT_GREEN
        else:
            row_bg = WHITE

        values = [
            channel,
            curr,
            rec,
            f"={get_column_letter(3)}{r}-{get_column_letter(2)}{r}",   # Δ Spend formula
            roas_val,
            f"={get_column_letter(2)}{r}*{get_column_letter(5)}{r}",   # Current Revenue
            f"={get_column_letter(3)}{r}*{get_column_letter(5)}{r}",   # Projected Revenue
            f"={get_column_letter(7)}{r}-{get_column_letter(6)}{r}",   # Revenue Δ
        ]

        for j, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.fill = _fill(row_bg)
            cell.border = _border()
            cell.alignment = _center() if j > 1 else _left()
            # Blue for hardcoded inputs
            if j in [1, 2, 3, 5]:
                cell.font = _font(color="0000FF", size=10)
            else:
                cell.font = _font(color="000000", size=10)
            # Format numbers
            if j in [2, 3, 4, 6, 7, 8]:
                cell.number_format = '#,##0'
            if j == 5:
                cell.number_format = '0.00"x"'
        ws.row_dimensions[r].height = 20

    # Totals row — no merging to avoid MergedCell conflict
    total_row = 5 + len(budget_df)
    for j in range(1, 9):
        cell = ws.cell(row=total_row, column=j)
        cell.font = _font(bold=True, color=WHITE, size=11)
        cell.fill = _fill(DARK_BLUE)
        cell.alignment = _center()
        cell.border = _border()

    ws.cell(row=total_row, column=1, value="TOTAL / PROJECTED GAIN")
    ws.cell(row=total_row, column=1).alignment = _left()

    for j in [2, 3, 6, 7, 8]:
        col = get_column_letter(j)
        cell = ws.cell(row=total_row, column=j,
                       value=f"=SUM({col}5:{col}{total_row-1})")
        cell.number_format = '#,##0'
    ws.row_dimensions[total_row].height = 24

    # Column widths
    widths = [22, 20, 20, 16, 12, 22, 22, 18]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, get_column_letter(i), w)


# ── SHEET 3: Incrementality Results ────────────────────────

def build_incrementality_sheet(wb, inc_df):
    ws = wb.create_sheet("Incrementality Results")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = "INCREMENTALITY TEST RESULTS — 70/30 Holdout Split"
    ws["A1"].font = _font(bold=True, color=WHITE, size=13)
    ws["A1"].fill = _fill(DARK_BLUE)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 32

    cols = ["channel", "test_n", "holdout_n", "test_purchase_rate",
            "holdout_purchase_rate", "incremental_rate", "p_value",
            "significant_95", "incremental_roas", "verdict"]
    headers = ["Channel", "Test N", "Holdout N", "Test CVR",
               "Holdout CVR", "Incremental Rate", "P-Value",
               "Significant (95%)", "Incremental ROAS", "Verdict"]

    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=j, value=h)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.fill = _fill(MID_BLUE)
        cell.alignment = _center()
        cell.border = _border()
    ws.row_dimensions[3].height = 28

    for r, (_, row) in enumerate(inc_df.iterrows(), 4):
        roas = row["incremental_roas"]
        bg = LIGHT_RED if roas < 0 else (YELLOW_BG if roas < 1 else (LIGHT_GREEN if roas >= 3 else WHITE))

        for j, col in enumerate(cols, 1):
            val = row[col]
            cell = ws.cell(row=r, column=j, value=val)
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.font = _font(size=10)
            cell.alignment = _center() if j > 1 else _left()
            if col in ["test_purchase_rate", "holdout_purchase_rate", "incremental_rate"]:
                cell.number_format = "0.00%"
            elif col == "incremental_roas":
                cell.number_format = '0.00"x"'
            elif col in ["test_n", "holdout_n"]:
                cell.number_format = "#,##0"
        ws.row_dimensions[r].height = 18

    # Legend
    leg_row = 4 + len(inc_df) + 2
    ws.cell(row=leg_row, column=1, value="LEGEND:").font = _font(bold=True, size=10)
    legend = [
        (LIGHT_GREEN, "Strong — ROAS ≥ 3x → Increase budget"),
        (WHITE,       "Positive — ROAS 1–3x → Maintain"),
        (YELLOW_BG,   "Caution — ROAS < 1x → Reduce budget"),
        (LIGHT_RED,   "Danger — Negative ROAS → Stop spending"),
    ]
    for i, (color, label) in enumerate(legend, leg_row + 1):
        ws.cell(row=i, column=1, value="").fill = _fill(color)
        ws.cell(row=i, column=1).border = _border()
        ws.cell(row=i, column=2, value=label).font = _font(size=10)
        ws.cell(row=i, column=2).alignment = _left()

    widths = [22, 12, 12, 12, 14, 16, 10, 16, 16, 45]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, get_column_letter(i), w)


# ── SHEET 4: Attribution Comparison ────────────────────────

def build_attribution_sheet(wb, attr_path):
    ws = wb.create_sheet("Attribution Models")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "MULTI-TOUCH ATTRIBUTION MODEL COMPARISON"
    ws["A1"].font = _font(bold=True, color=WHITE, size=13)
    ws["A1"].fill = _fill(DARK_BLUE)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    ws["A2"] = ("Higher 'Disagreement' score = models strongly disagree = "
                "don't trust any single model for that channel")
    ws["A2"].font = _font(italic=True, color=DARK_GRAY, size=9)
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 16

    attr_df = pd.read_csv(attr_path, index_col=0)
    headers = ["Channel"] + list(attr_df.columns)
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=j, value=h)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.fill = _fill(MID_BLUE)
        cell.alignment = _center()
        cell.border = _border()
    ws.row_dimensions[4].height = 28

    for r, (channel, row) in enumerate(attr_df.iterrows(), 5):
        disagreement = row.get("Disagreement", 0)
        bg = LIGHT_RED if disagreement > 0.10 else (YELLOW_BG if disagreement > 0.05 else WHITE)

        ws.cell(row=r, column=1, value=channel).font = _font(bold=True, size=10)
        ws.cell(row=r, column=1).fill = _fill(bg)
        ws.cell(row=r, column=1).border = _border()
        ws.cell(row=r, column=1).alignment = _left()

        for j, val in enumerate(row.values, 2):
            cell = ws.cell(row=r, column=j, value=val)
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.font = _font(size=10)
            cell.alignment = _center()
            if j < len(headers):   # percentage columns
                cell.number_format = "0.00%"
            else:                  # disagreement column
                cell.number_format = "0.0000"
        ws.row_dimensions[r].height = 18

    widths = [22, 14, 14, 12, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, get_column_letter(i), w)


# ── MAIN ───────────────────────────────────────────────────

def generate_excel_report(
    incrementality_csv: str = "data/outputs/incrementality_results.csv",
    budget_csv:         str = "data/outputs/budget_optimisation.csv",
    attribution_csv:    str = "data/outputs/attribution_comparison.csv",
    output_path:        str = "reports/budget_recommendation.xlsx",
):
    inc_df    = pd.read_csv(incrementality_csv)
    budget_df = pd.read_csv(budget_csv)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_executive_summary(wb, inc_df, budget_df)
    build_budget_sheet(wb, budget_df, inc_df)
    build_incrementality_sheet(wb, inc_df)
    build_attribution_sheet(wb, attribution_csv)

    os.makedirs("reports", exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel report saved → {output_path}")
    print(f"   Sheets: Executive Summary | Budget Recommendation | Incrementality Results | Attribution Models")
    return output_path


if __name__ == "__main__":
    generate_excel_report()